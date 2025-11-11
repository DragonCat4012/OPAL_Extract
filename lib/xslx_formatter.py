import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter

html_colors = [
    "#85E6DF",  # Soft Green
    "#55A6DB",  # Bright Blue
    "#897CEE",  # Violet
    "#BE7DDA",  # Purple
    "#F7AC80",  # Red-Orange
    "#F56E6E",  # Light Red
    "#D374A0",  # Pink
    "#EE8F64",  # Carrot Orange
    "#E7D27B",  # Yellow
    "#79F38F",  # Green
    "#5EC2AE",  # Dark Turquoise
    "#44B8CC",  # Turquoise
    "#8599F3",  # Blue
    "#75C1F3",  # Bright Blue
    "#C279DF",  # Amethyst
    "#56E090",  # Emerald
    "#FAC165",  # Orange
    "#7F8C8D",  # Grey
    "#DB6254",  # Strong Red
    "#FFB142",  # Golden Yellow
    "#F39C12",  # Orange
    "#F8C291",  # Peach
    "#FFE156",  # Bright Yellow
    "#FF7952",  # Coral
    "#6AB04C",  # Light Green
    "#12CBC4",  # Dark Turquoise
    "#3C6382",  # Dark Blue
    "#B8E1FF",  # Light Blue
]


def format_xlsx(input_file_path, group, all_groups):
    """Format XLSX Sheet and color in rows where dropboxes are there"""
    # TODO: maybe only dropboxes with content in them?

    output_file_path = os.path.join(
        input_file_path.split(os.sep)[0], f"Bewertung_{group}.xlsx"
    )

    columns_to_remove = [
        "Anrede",
        "Studiengruppe",
        "Organisationseinheit",
        "Fachsemester",
        "Studiengang",
        "Studienabschluss",
        "Institution",
        "Standort",
        "Startdatum",
        "Dauer",
        "Versuche"
    ]

    _edit_colums_with_group_info(
        input_file_path, output_file_path, columns_to_remove, all_groups
    )

    move_rows_by_color(output_file_path)


def _edit_colums_with_group_info(
    input_file_path, output_file_path, columns_to_remove, all_groups
):
    df = pd.read_excel(input_file_path)
    df.drop(columns=columns_to_remove, inplace=True, errors="ignore")
    df.to_excel(output_file_path, engine="openpyxl", index=False)

    file_imma_nr_list = df["Matrikelnummer"].astype(str).tolist()

    workbook = load_workbook(output_file_path)
    sheet = workbook.active

    _add_missing_headers(sheet)

    color_index = 0

    # assign colors
    for t in all_groups:
        t.color = html_colors[color_index].replace("#", "")
        color_index += 1

    header = df.columns.tolist()
    target_index = header.index("Matrikelnummer")

    # tmp save newly added rows with their color
    new_added_row_colors = {}
    insert_empty_row_before_other_group_members = True

    for row in sheet.iter_rows(min_row=1):
        value: str = f"{row[target_index].value}"

        for t in all_groups:
            if t.is_member(value):
                # fill rows with same color
                fill = PatternFill(
                    start_color=t.color, end_color=t.color, fill_type="solid"
                )
                for cell in row:
                    cell.fill = fill

            # Add member form other groups below
            for imma_nr in t.member:
                if not imma_nr in file_imma_nr_list:
                    file_imma_nr_list.append(imma_nr)

                    new_added_row_colors[imma_nr] = t.color

                    new_row_data = {
                        "Matrikelnummer": imma_nr,
                        "AndereGruppe": "OtherGroup",
                        "Punkte": 0,
                    }

                    if insert_empty_row_before_other_group_members:
                        df = df._append({"Matrikelnummer": ""}, ignore_index=True)
                        insert_empty_row_before_other_group_members = False

                    df = df._append(new_row_data, ignore_index=True)

                    for index, row in df.iterrows():
                        for col_index, value in enumerate(row):
                            sheet.cell(row=index + 2, column=col_index + 1, value=value)

    # Update row colors of other gorup cells
    for row in sheet.iter_rows(min_row=1):
        cell_A = row[:1][0]
        cell_A.alignment = Alignment(horizontal="center")

        value: str = f"{row[target_index].value}"
        if value in new_added_row_colors:
            color = new_added_row_colors[value]
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for cell in row:
                cell.fill = fill
                color_index += 1

    # Fix colum width
    dim_holder = DimensionHolder(worksheet=sheet)

    for col in range(sheet.min_column, sheet.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(
            sheet, min=col, max=col, width=20
        )
    sheet.column_dimensions = dim_holder
    workbook.save(output_file_path)
    # print(f"{count} People marked in Excel sheet")


def _add_missing_headers(sheet):
    arr = ["AndereGruppe", "Punkte", "Kommentar"]
    headers = [cell.value for cell in sheet[1]]

    h1 = len(headers) + 1
    for val in arr:
        if val not in headers:
            sheet.cell(row=1, column=h1, value=val)
            h1 += 1

def move_rows_by_color(filename):
    """Reorder rows based on their color"""
    print("Reorder rows")
    wb = load_workbook(filename)
    ws = wb.active

    white_rows = []

    # Group rows based on color
    color_dict = {}
    for row in ws.iter_rows(min_row=2):  # Skip header row if there is one
        color = get_color(row)

        if color == '00000000':  # ungraded people
            white_rows.append(row)
            continue

        if color not in color_dict:
            color_dict[color] = []
        color_dict[color].append(row)

    # Clear the existing rows except header
    ws.delete_rows(2, ws.max_row) 

    # Re-add rows sorted by color
    start_row = 2  # Start writing after header
    for color, rows in color_dict.items():
        for row in rows:
            for cell in row:
                new_cell = ws.cell(row=start_row, column=cell.col_idx, value=cell.value)
                new_cell.fill = PatternFill(start_color=cell.fill.start_color.index,
                                             end_color=cell.fill.end_color.index,
                                             fill_type=cell.fill.fill_type)
                new_cell.alignment = Alignment(horizontal="center")
            start_row += 1

    # Add spacer between groups and non-groups
    if white_rows:
        start_row += 1  # Create an empty row before writing white rows

    # Re-add empty rows
    for row in white_rows:
        for cell in row:
            new_cell = ws.cell(row=start_row, column=cell.col_idx, value=cell.value)
            new_cell.alignment = Alignment(horizontal="center")
        start_row += 1

    wb.save(filename)

def get_color(row):
    """Get row group color"""
    cell = row[0]
    return cell.fill.start_color.index if cell.fill else None
