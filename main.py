import zipfile
import os
import sys
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re

from team import Team

dropbox_immanrs = [] 
imma_nr_map = {} # read entrys from readme.txt?
groups = []

html_colors = [
    "#FF5733",  # Red-Orange
    "#33FF57",  # Green
    "#3357FF",  # Blue
    "#F1C40F",  # Yellow
    "#8E44AD",  # Purple
    "#E67E22",  # Carrot Orange
    "#3498DB",  # Bright Blue
    "#1ABC9C",  # Turquoise
    "#E74C3C",  # Red
    "#9B59B6",  # Amethyst
    "#2ECC71",  # Emerald
    "#F39C12",  # Orange
    "#D35400",  # Pumpkin
    "#7F8C8D",  # Grey
    "#C0392B",  # Strong Red
    "#2980B9",  # Bright Blue
    "#16A085",  # Dark Turquoise
    "#FFB142",  # Golden Yellow
    "#F39C12",  # Orange
    "#6C5CE7",  # Violet
    "#F8C291",  # Peach
    "#FF6B6B",  # Light Red
    "#4ECDC4",  # Soft Green
    "#FFE156",  # Bright Yellow
    "#FF7952",  # Coral
    "#6AB04C",  # Light Green
    "#12CBC4",  # Dark Turquoise
    "#B33771",  # Pink
    "#3C6382",  # Dark Blue
    "#B8E1FF",  # Light Blue
]

def extract_zip(zip_file_path, extraction_folder, print_info = False):
    os.makedirs(extraction_folder, exist_ok=True)

    with zipfile.ZipFile(zip_file_path, "r") as zip_ref:
        zip_ref.extractall(extraction_folder)

    #print(f"> Successfully extracted to {extraction_folder}")

    rating_file_name = None
    for item in os.listdir(extraction_folder):
        if item.endswith(".xlsx"):
            print(f"\t{item}")
            rating_file_name = os.path.join(extraction_folder, item)
    move_extracted_content(extraction_folder, print_info)

    #print(f"> Returning : {rating_file_name}")
    return rating_file_name


def add_team(members, imma_nr):
    for t in groups:
        if t.is_member(imma_nr):
            return 
    members.append(imma_nr)
    groups.append(Team(members))

def move_extracted_content(parent_folder, print_info):
    """Move dropboxes and extratc content from them"""
    global dropbox_immanrs
    extracted_dirs = [
        d
        for d in os.listdir(parent_folder)
        if os.path.isdir(os.path.join(parent_folder, d))
    ]

    if not extracted_dirs:
        #print("No folders found to extract.")
        return

    nested_folder = os.path.join(parent_folder, extracted_dirs[0])

    if print_info:
        print("---------------------Dropboxes-------------------------")

    for item in os.listdir(nested_folder):
        if print_info:
            print(f"\t{item}")
        imma_nr = f"{item}".split("_")[-1]
        dropbox_immanrs.append(imma_nr)

        item_path = os.path.join(nested_folder, item)

        # Assignment entzip
        if os.path.isdir(item_path):  
            all_files = os.listdir(item_path)
            for file in all_files:
                file_path = os.path.join(item_path, file)
               
                if file_path.endswith(".zip"):
                    print(f"\t\tAssignment: {file}")
                    extract_zip(file_path, item_path)

                    all_files2 = os.listdir(item_path)
                    for file2 in all_files2:
                        if f"{file2}".lower() == "readme.txt":
                            print(f"\t\t\t{file2}")
                            with open(os.path.join(item_path, file2), 'r') as file:
                                content = file.read()
                                seven_digit_numbers = re.findall(r'\b\d{7}\b', content)
                                if imma_nr in seven_digit_numbers:
                                    seven_digit_numbers.remove(imma_nr)
                                print(f"\t\t\tGroup members: {seven_digit_numbers}") 
                                add_team(seven_digit_numbers, imma_nr)                      

        # Nested Zips
        if item.endswith(".zip"):
            extract_zip(item_path, nested_folder, True)

            for extracted_item in os.listdir(nested_folder):
                print(f"\t\t{extracted_item}")
                
                extracted_item_path = os.path.join(nested_folder, extracted_item)

                if not "dropboxes" in extracted_item_path:
                    continue
                print("> Dropboxes:" + extracted_item_path)

                if os.path.isdir(extracted_item_path):
                    for file in os.listdir(extracted_item_path):
                        shutil.move(
                            os.path.join(extracted_item_path, file), parent_folder
                        )
                    os.rmdir(extracted_item_path)

def formatXSL(input_file_path, group):
    """Format XLSX Sheet and color in rows where dropboxes are there"""
    # TODO: maybe only dropboxes with content in them?

    output_file_path = os.path.join(input_file_path.split(os.sep)[0], f"Bewertung_{group}.xlsx"
    )

    columns_to_remove = [
        "Anrede",
        "Studiengruppe",
        "Organisationseinheit",
        "Fachsemester",
        "Studiengang",
        "Studienabschluss",
        "Institution",
        "Standort",
        "Startdatum",
        "Dauer",
    ]

    remove_columns_from_xls(input_file_path, output_file_path, columns_to_remove)

def remove_columns_from_xls(input_file_path, output_file_path, columns_to_remove):
    df = pd.read_excel(input_file_path)
    df.drop(columns=columns_to_remove, inplace=True, errors='ignore')
    df.to_excel(output_file_path, engine="openpyxl", index=False)

    file_imma_nr_list = df["Matrikelnummer"].astype(str).tolist()

    workbook = load_workbook(output_file_path)
    sheet = workbook.active

    color_index = 0
    
    # assign colors
    for t in groups:
        t.color = html_colors[color_index].replace("#", '')
        color_index += 1
    
    header = df.columns.tolist()
    target_index = header.index("Matrikelnummer")
    
    # tmp save newly added rows with their color
    new_added_row_colors = {}

    for row in sheet.iter_rows(min_row=1):
        value: str = f"{row[target_index].value}"

        for t in groups:
            if t.is_member(value):
                # fill rows with same color
                fill = PatternFill(start_color=t.color, end_color=t.color, fill_type="solid")
                for cell in row:
                    cell.fill = fill
                    

             # Add member form other groups below
            for imma_nr in t.member:
                if not imma_nr in file_imma_nr_list:
                    file_imma_nr_list.append(imma_nr)

                    new_added_row_colors[imma_nr] = t.color
           
                    new_row_data = {
                        "Matrikelnummer": imma_nr,
                        "OtherGroup": "OtherGroup"
                    }
                    df = df._append(new_row_data, ignore_index=True)

                    for index, row in df.iterrows():
                        for col_index, value in enumerate(row):
                            sheet.cell(row=index + 2, column=col_index + 1, value=value)
    

    # Update row colors of other gorup cells
    for row in sheet.iter_rows(min_row=1):
        value: str = f"{row[target_index].value}"
        if value in new_added_row_colors:
            color = new_added_row_colors[value]
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for cell in row:
                cell.fill = fill
                color_index += 1

    workbook.save(output_file_path)
    #print(f"{count} People marked in Excel sheet")

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python script.py <zip_file_path> <groupnumber>")
        sys.exit(1)

    zip_file_path = sys.argv[1]
    groupnumber = sys.argv[2]
    foldername = f"GruMCI G{groupnumber}"

    # rmeove dir if already exists
    if os.path.exists(foldername):
        shutil.rmtree(foldername)
        print("Directory ZIP1 has been deleted successfully.")

    filePath = extract_zip(zip_file_path, foldername)
    formatXSL(filePath, groupnumber)
